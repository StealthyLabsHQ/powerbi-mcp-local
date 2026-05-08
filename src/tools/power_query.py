"""Power Query (M language) tools for the Power BI MCP server."""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any

from m_expression_security import validate_m_expression as validate_m_expression_policy
from pbi_connection import (
    PowerBIConfigurationError,
    PowerBINotFoundError,
    PowerBIValidationError,
    find_named,
    normalize_token,
    ok,
    serialize_value,
)
from security import (
    inspect_excel_archive,
    redact_sensitive_data,
    resolve_local_path,
    validate_expression_text,
    validate_model_object_name,
)


def _m_string(value: str) -> str:
    return '"' + str(value).replace('"', '""') + '"'

def _validate_m_expression(expression: str) -> None:
    validate_expression_text(expression)
    validate_m_expression_policy(expression, allow_external=os.environ.get("PBI_MCP_ALLOW_EXTERNAL_M", "0") == "1")


def _source_type_token(partition: Any) -> str:
    source = getattr(partition, "Source", None)
    source_name = type(source).__name__ if source is not None else ""
    raw = str(getattr(partition, "SourceType") or "") if hasattr(partition, "SourceType") else ""
    token = normalize_token(source_name or raw)
    if token in {"m", "none", "query", "entity", "calculated", "calculationgroup"}:
        return {
            "m": "m",
            "none": "none",
            "query": "query",
            "entity": "entity",
            "calculated": "calculated",
            "calculationgroup": "calculation_group",
        }[token]
    if token.endswith("mpartitionsource"):
        return "m"
    if token.endswith("calculatedpartitionsource"):
        return "calculated"
    if token.endswith("querypartitionsource"):
        return "query"
    if token.endswith("entitypartitionsource"):
        return "entity"
    if token.endswith("policyrangepartitionsource") or token == "policyrange":
        return "policy_range"
    return "unknown"


def _partition_expression(partition: Any) -> str:
    source = getattr(partition, "Source", None)
    token = _source_type_token(partition)
    if source is not None:
        if token in {"m", "calculated", "policy_range"} and hasattr(source, "Expression"):
            return str(source.Expression or "")
        if token == "query" and hasattr(source, "Query"):
            return str(source.Query or "")
        if hasattr(source, "Expression"):
            return str(source.Expression or "")
    if hasattr(partition, "Expression"):
        return str(getattr(partition, "Expression") or "")
    return ""


def _partition_payload(table: Any, partition: Any) -> dict[str, Any]:
    expression = _partition_expression(partition)
    return {
        "table": str(table.Name),
        "partition": str(partition.Name),
        "source_type": _source_type_token(partition),
        "source_type_raw": serialize_value(getattr(partition, "SourceType", None)),
        "m_expression": redact_sensitive_data(expression),
        "expression_length": len(expression),
    }


def _get_target_partition(model: Any, table_name: str, partition_name: str | None = None) -> tuple[Any, Any]:
    table = find_named(model.Tables, table_name)
    if table is None:
        raise PowerBINotFoundError(f"Table '{table_name}' was not found.", details={"table": table_name})
    count = int(table.Partitions.Count)
    if count == 0:
        raise PowerBINotFoundError(f"Table '{table_name}' has no partitions.", details={"table": table_name})
    if partition_name:
        partition = find_named(table.Partitions, partition_name)
        if partition is None:
            raise PowerBINotFoundError(
                f"Partition '{partition_name}' was not found on table '{table_name}'.",
                details={"table": table_name, "partition": partition_name},
            )
        return table, partition
    if count > 1:
        raise PowerBIValidationError(
            f"Table '{table_name}' has multiple partitions. Specify partition_name explicitly.",
            details={"table": table_name, "partitions": [str(item.Name) for item in table.Partitions]},
        )
    # Avoid int indexing on the .NET collection (which expects a String name);
    # iterate instead and return the first partition.
    for partition in table.Partitions:
        return table, partition
    raise PowerBINotFoundError(f"Table '{table_name}' has no partitions.", details={"table": table_name})


def _ensure_m_supported(database: Any) -> None:
    compatibility = getattr(database, "CompatibilityLevel", None)
    if compatibility is not None and int(compatibility) < 1400:
        raise PowerBIValidationError(
            "Power Query partition injection requires compatibility level 1400 or higher.",
            details={"compatibility_level": compatibility},
        )


def _set_partition_m_expression(manager: Any, database: Any, partition: Any, expression: str) -> str:
    _ensure_m_supported(database)
    source_type = _source_type_token(partition)
    if source_type == "calculated":
        raise PowerBIValidationError(
            f"Partition '{partition.Name}' is calculated and cannot be overwritten with an M expression.",
            details={"partition": str(partition.Name), "source_type": source_type},
        )
    source = getattr(partition, "Source", None)
    if source_type != "m" or source is None or not hasattr(source, "Expression"):
        if not hasattr(manager.tom, "MPartitionSource"):
            raise PowerBIConfigurationError("This TOM build does not expose MPartitionSource.")
        source = manager.tom.MPartitionSource()
        partition.Source = source
    source.Expression = expression
    return _source_type_token(partition)


def _request_refresh(manager: Any, table: Any, refresh_after: bool) -> None:
    if refresh_after and hasattr(manager.tom, "RefreshType"):
        table.RequestRefresh(manager.tom.RefreshType.Full)


def _load_excel_sheet_names(excel_path: str) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise PowerBIConfigurationError("openpyxl is required for Excel import query helpers.") from exc
    path = inspect_excel_archive(excel_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()


def _ensure_file(path_value: str, *, kind: str, allowed_extensions: set[str] | None = None) -> str:
    path = resolve_local_path(path_value, must_exist=True, allowed_extensions=allowed_extensions)
    if not path.is_file():
        raise PowerBINotFoundError(f"{kind} '{path}' was not found.", details={"path": str(path)})
    return str(path)


def _ensure_folder(path_value: str) -> str:
    path = resolve_local_path(path_value, must_exist=True)
    if not path.is_dir():
        raise PowerBINotFoundError(f"Folder '{path}' was not found.", details={"path": str(path)})
    return str(path)


def _build_excel_m(excel_path: str, sheet_name: str, promote_headers: bool = True) -> str:
    final_step = "Promoted" if promote_headers else "Sheet"
    steps = [
        f"    Source = Excel.Workbook(File.Contents({_m_string(excel_path)}), null, true)",
        f"    Sheet = Source{{[Item={_m_string(sheet_name)},Kind=\"Sheet\"]}}[Data]",
    ]
    if promote_headers:
        steps.append("    Promoted = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true])")
    return "let\n" + ",\n".join(steps) + f"\nin\n    {final_step}"


def _build_csv_m(
    csv_path: str,
    *,
    delimiter: str = ",",
    encoding: int = 65001,
    quote_style: str = "csv",
    promote_headers: bool = True,
) -> str:
    token = normalize_token(quote_style)
    if token not in {"csv", "none"}:
        raise PowerBIValidationError(
            "quote_style must be 'csv' or 'none'.",
            details={"quote_style": quote_style},
        )
    final_step = "Promoted" if promote_headers else "Source"
    # Each top-level let step is a single string; steps are joined by ",\n".
    # The Csv.Document call spans multiple lines INSIDE one step (no comma between its lines).
    csv_doc_step = (
        "    Source = Csv.Document(\n"
        f"        File.Contents({_m_string(csv_path)}),\n"
        f"        [Delimiter={_m_string(delimiter)}, Encoding={encoding}, QuoteStyle=QuoteStyle.{token.title()}]\n"
        "    )"
    )
    steps = [csv_doc_step]
    if promote_headers:
        steps.append("    Promoted = Table.PromoteHeaders(Source, [PromoteAllScalars=true])")
    return "let\n" + ",\n".join(steps) + f"\nin\n    {final_step}"


def _build_folder_m(
    folder_path: str,
    *,
    extension_filter: str | None = None,
    include_hidden_files: bool = False,
) -> str:
    final_step = "FilteredExtension" if extension_filter else "VisibleFiles"
    extension = extension_filter if not extension_filter or extension_filter.startswith(".") else "." + extension_filter
    steps = [f"    Source = Folder.Files({_m_string(folder_path)})"]
    if include_hidden_files:
        steps.append("    VisibleFiles = Source")
    else:
        steps.append("    VisibleFiles = Table.SelectRows(Source, each [Attributes]?[Hidden]? <> true)")
    if extension:
        steps.append(
            f"    FilteredExtension = Table.SelectRows(VisibleFiles, each Text.Lower([Extension]) = {_m_string(extension.lower())})"
        )
    return "let\n" + ",\n".join(steps) + f"\nin\n    {final_step}"


def _build_auto_sheet_map(model: Any, sheet_names: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for sheet_name in sheet_names:
        table = find_named(model.Tables, sheet_name)
        if table is None or bool(getattr(table, "IsHidden", False)):
            continue
        mapping[sheet_name] = str(table.Name)
    return mapping


def pbi_get_power_query_tool(manager: Any, *, table: str, partition_name: str | None = None) -> dict[str, Any]:
    """Read the M expression for a specific table partition."""
    validate_model_object_name(table)
    if partition_name:
        validate_model_object_name(partition_name)

    def _reader(state: Any) -> dict[str, Any]:
        tbl, partition = _get_target_partition(state.database.Model, table, partition_name)
        return {"query": _partition_payload(tbl, partition), "connection": state.snapshot()}

    payload = manager.run_read("get_power_query", _reader)
    return ok(
        f"Power Query expression retrieved for table '{table}'.",
        query=payload["query"],
        connection=payload["connection"],
    )


def pbi_list_power_queries_tool(manager: Any, *, include_hidden: bool = False) -> dict[str, Any]:
    """List table partitions and their current source expressions."""

    def _reader(state: Any) -> dict[str, Any]:
        queries = []
        for table in state.database.Model.Tables:
            is_hidden = bool(getattr(table, "IsHidden", False))
            if is_hidden and not include_hidden:
                continue
            partitions = [_partition_payload(table, partition) for partition in table.Partitions]
            queries.append(
                {
                    "table": str(table.Name),
                    "is_hidden": is_hidden,
                    "partition_count": len(partitions),
                    "partitions": partitions,
                }
            )
        queries.sort(key=lambda item: item["table"].casefold())
        return {"queries": queries, "connection": state.snapshot()}

    payload = manager.run_read("list_power_queries", _reader)
    return ok(
        "Power Query expressions listed successfully.",
        queries=payload["queries"],
        connection=payload["connection"],
    )


def pbi_set_power_query_tool(
    manager: Any,
    *,
    table: str,
    m_expression: str,
    partition_name: str | None = None,
    refresh_after: bool = False,
) -> dict[str, Any]:
    """Write or update an M expression on a table partition."""
    validate_model_object_name(table)
    if partition_name:
        validate_model_object_name(partition_name)
    _validate_m_expression(m_expression)

    def _mutator(state: Any, database: Any, model: Any) -> dict[str, Any]:
        tbl, partition = _get_target_partition(model, table, partition_name)
        previous = _partition_payload(tbl, partition)
        new_source_type = _set_partition_m_expression(manager, database, partition, m_expression)
        _request_refresh(manager, tbl, refresh_after)
        return {
            "query": {
                "table": str(tbl.Name),
                "partition": str(partition.Name),
                "previous_source_type": previous["source_type"],
                "source_type": new_source_type,
                "previous_expression_length": previous["expression_length"],
                "expression_length": len(m_expression),
                "m_expression": redact_sensitive_data(m_expression),
                "refresh_requested": refresh_after,
            }
        }

    payload = manager.execute_write("set_power_query", _mutator)
    return ok(
        f"Power Query expression updated for table '{table}'.",
        query=payload["query"],
        save_result=payload["save_result"],
        persistence=payload.get("persistence"),
        connection=payload["connection"],
    )


def pbi_create_import_query_tool(
    manager: Any,
    *,
    table: str,
    excel_path: str,
    sheet_name: str,
    partition_name: str | None = None,
    promote_headers: bool = True,
    refresh_after: bool = True,
) -> dict[str, Any]:
    """Generate and inject an Excel import Power Query for a table."""
    validate_model_object_name(table)
    if partition_name:
        validate_model_object_name(partition_name)
    workbook_path = str(inspect_excel_archive(excel_path))
    available_sheets = _load_excel_sheet_names(workbook_path)
    if sheet_name not in available_sheets:
        raise PowerBINotFoundError(
            f"Sheet '{sheet_name}' was not found in workbook '{workbook_path}'.",
            details={"path": workbook_path, "sheet": sheet_name, "available_sheets": available_sheets},
        )
    m_expression = _build_excel_m(workbook_path, sheet_name, promote_headers)
    response = pbi_set_power_query_tool(
        manager,
        table=table,
        m_expression=m_expression,
        partition_name=partition_name,
        refresh_after=refresh_after,
    )
    if response.get("ok"):
        response["message"] = f"Excel import query created for table '{table}' from sheet '{sheet_name}'."
    return response


def pbi_create_csv_import_query_tool(
    manager: Any,
    *,
    table: str,
    csv_path: str,
    partition_name: str | None = None,
    delimiter: str = ",",
    encoding: int = 65001,
    quote_style: str = "csv",
    promote_headers: bool = True,
    refresh_after: bool = True,
) -> dict[str, Any]:
    """Generate and inject a CSV import Power Query for a table."""
    validate_model_object_name(table)
    if partition_name:
        validate_model_object_name(partition_name)
    csv_file = _ensure_file(csv_path, kind="CSV file", allowed_extensions={".csv", ".txt"})
    m_expression = _build_csv_m(
        csv_file,
        delimiter=delimiter,
        encoding=encoding,
        quote_style=quote_style,
        promote_headers=promote_headers,
    )
    response = pbi_set_power_query_tool(
        manager,
        table=table,
        m_expression=m_expression,
        partition_name=partition_name,
        refresh_after=refresh_after,
    )
    if response.get("ok"):
        response["message"] = f"CSV import query created for table '{table}' from '{csv_file}'."
    return response


def pbi_create_folder_import_query_tool(
    manager: Any,
    *,
    table: str,
    folder_path: str,
    partition_name: str | None = None,
    extension_filter: str | None = None,
    include_hidden_files: bool = False,
    refresh_after: bool = True,
) -> dict[str, Any]:
    """Generate and inject a folder import Power Query for a table."""
    validate_model_object_name(table)
    if partition_name:
        validate_model_object_name(partition_name)
    folder = _ensure_folder(folder_path)
    m_expression = _build_folder_m(
        folder,
        extension_filter=extension_filter,
        include_hidden_files=include_hidden_files,
    )
    response = pbi_set_power_query_tool(
        manager,
        table=table,
        m_expression=m_expression,
        partition_name=partition_name,
        refresh_after=refresh_after,
    )
    if response.get("ok"):
        response["message"] = f"Folder import query created for table '{table}' from '{folder}'."
    return response


def pbi_bulk_import_excel_tool(
    manager: Any,
    *,
    excel_path: str,
    sheet_table_map: dict[str, str] | None = None,
    promote_headers: bool = True,
    refresh_after: bool = True,
) -> dict[str, Any]:
    """Bulk-create Excel import queries for multiple tables."""
    workbook_path = str(inspect_excel_archive(excel_path))
    available_sheets = _load_excel_sheet_names(workbook_path)

    def _mutator(state: Any, database: Any, model: Any) -> dict[str, Any]:
        mapping = dict(sheet_table_map or _build_auto_sheet_map(model, available_sheets))
        available_sheet_set = set(available_sheets)
        results = []
        created = 0
        for sheet_name, table_name in mapping.items():
            table = find_named(model.Tables, table_name)
            if sheet_name not in available_sheet_set:
                results.append({"table": table_name, "sheet": sheet_name, "status": "skipped", "reason": "sheet_not_found"})
                continue
            if table is None:
                results.append({"table": table_name, "sheet": sheet_name, "status": "skipped", "reason": "table_not_found"})
                continue
            if bool(getattr(table, "IsHidden", False)):
                results.append({"table": table_name, "sheet": sheet_name, "status": "skipped", "reason": "table_hidden"})
                continue
            if int(table.Partitions.Count) == 0:
                results.append({"table": table_name, "sheet": sheet_name, "status": "skipped", "reason": "no_partitions"})
                continue
            if int(table.Partitions.Count) > 1:
                results.append(
                    {
                        "table": table_name,
                        "sheet": sheet_name,
                        "status": "skipped",
                        "reason": "multiple_partitions",
                        "partitions": [str(item.Name) for item in table.Partitions],
                    }
                )
                continue
            partition = next(iter(table.Partitions))
            try:
                m_expression = _build_excel_m(workbook_path, sheet_name, promote_headers)
                _validate_m_expression(m_expression)
                new_source_type = _set_partition_m_expression(
                    manager,
                    database,
                    partition,
                    m_expression,
                )
                _request_refresh(manager, table, refresh_after)
            except Exception as exc:
                results.append(
                    {
                        "table": table_name,
                        "sheet": sheet_name,
                        "status": "skipped",
                        "reason": getattr(exc, "message", str(exc)),
                        "error_code": getattr(exc, "code", "internal_error"),
                    }
                )
                continue
            created += 1
            results.append(
                {
                    "table": str(table.Name),
                    "sheet": sheet_name,
                    "partition": str(partition.Name),
                    "status": "created",
                    "source_type": new_source_type,
                }
            )
        return {
            "excel_path": workbook_path,
            "sheet_table_map": mapping,
            "results": results,
            "created": created,
            "skipped": len(results) - created,
            "refresh_requested": refresh_after,
        }

    payload = manager.execute_write("bulk_import_excel", _mutator)
    return ok(
        f"Bulk import done: {payload['created']} queries created, {payload['skipped']} skipped.",
        excel_path=payload["excel_path"],
        sheet_table_map=payload["sheet_table_map"],
        results=payload["results"],
        created=payload["created"],
        skipped=payload["skipped"],
        refresh_requested=payload["refresh_requested"],
        save_result=payload["save_result"],
        persistence=payload.get("persistence"),
        connection=payload["connection"],
    )


def _m_identifier_safe(name: str) -> str:
    """Validate that a parameter name is a safe bare M identifier.

    Power Query parameters used as bare identifiers in M expressions must be
    plain alphanumeric/underscore names (no spaces, no leading digits).
    """
    if not name or not name[0].isalpha() and name[0] != "_":
        raise PowerBIValidationError(
            "parameter_name must start with a letter or underscore.",
            details={"parameter_name": name},
        )
    if not all(ch.isalnum() or ch == "_" for ch in name):
        raise PowerBIValidationError(
            "parameter_name may only contain letters, digits, and underscore.",
            details={"parameter_name": name},
        )
    return name


def pbi_parameterize_data_source_tool(
    manager: Any,
    *,
    parameter_name: str,
    file_path: str,
    partitions: list[str] | None = None,
    dry_run: bool = False,
    refresh_after: bool = False,
) -> dict[str, Any]:
    """Make data sources portable via a Power Query parameter.

    Creates (or updates) an M parameter named ``parameter_name`` whose default
    value is ``file_path``. Then rewrites every M partition that string-matches
    ``file_path`` so it calls ``File.Contents(<parameter_name>)`` instead of the
    hardcoded path. After the parameter is wired in, a collaborator can change
    the file location once via Power BI Desktop's *Transform data → Manage
    parameters* dialog without editing each query.

    ``partitions``: optional explicit list of "Table[/Partition]" to rewrite. By
    default every M-source partition that references ``file_path`` is rewritten.
    """
    _m_identifier_safe(parameter_name)
    if not file_path:
        raise PowerBIValidationError("file_path must be a non-empty string.")
    # Resolve the path against the security policy so we never persist a path
    # that the policy would refuse to read.
    resolved = resolve_local_path(file_path, must_exist=False)
    canonical_path = str(resolved)

    target_partitions: set[str] | None = None
    if partitions:
        target_partitions = set()
        for entry in partitions:
            if "/" in entry:
                tbl, part = entry.split("/", 1)
            else:
                tbl, part = entry, ""
            target_partitions.add(f"{tbl.strip()}::{part.strip()}")

    # M parameter expression: literal string + meta record marking it as a parameter.
    parameter_expression = (
        f'"{canonical_path}" meta [IsParameterQuery=true, '
        f'Type=type text, IsParameterQueryRequired=true]'
    )

    def _act(state: Any, database: Any, model: Any) -> dict[str, Any]:
        # Step 1: ensure the parameter (NamedExpression) exists / is up to date.
        named_expression_action = "noop"
        existing = None
        try:
            for expr in model.Expressions:
                if str(expr.Name) == parameter_name:
                    existing = expr
                    break
        except Exception:
            existing = None
        if existing is None:
            if not hasattr(manager.tom, "NamedExpression"):
                raise PowerBIConfigurationError(
                    "This TOM build does not expose NamedExpression; cannot create M parameter."
                )
            new_expr = manager.tom.NamedExpression()
            new_expr.Name = parameter_name
            new_expr.Expression = parameter_expression
            try:
                # Set Kind to M (= 1) if the enum is reachable; not strictly required for PBI.
                new_expr.Kind = manager.tom.ExpressionKind.M  # type: ignore[attr-defined]
            except Exception:
                pass
            if not dry_run:
                model.Expressions.Add(new_expr)
            named_expression_action = "created"
        else:
            if existing.Expression != parameter_expression:
                if not dry_run:
                    existing.Expression = parameter_expression
                named_expression_action = "updated"

        # Step 2: rewrite each M partition that mentions file_path.
        rewrites: list[dict[str, Any]] = []
        for table in model.Tables:
            for partition in table.Partitions:
                key = f"{str(table.Name)}::{str(partition.Name)}"
                if target_partitions is not None and key not in target_partitions:
                    continue
                source_type = _source_type_token(partition)
                if source_type != "m":
                    continue
                expression = _partition_expression(partition)
                if file_path not in expression:
                    continue
                # Replace the hardcoded literal with a bare identifier reference.
                # Patterns covered:  "<file_path>"  (most common)
                # We also handle the security-sensitive case where someone wrapped the
                # path in single quotes — but PBI's M only uses double quotes here.
                quoted = f'"{file_path}"'
                if quoted in expression:
                    new_expression = expression.replace(quoted, parameter_name)
                else:
                    # Fallback: substring replace that strips the surrounding quotes
                    # from the matched chunk.
                    new_expression = expression.replace(file_path, parameter_name)
                if new_expression == expression:
                    continue
                _validate_m_expression(new_expression)
                if not dry_run:
                    _set_partition_m_expression(manager, database, partition, new_expression)
                    _request_refresh(manager, table, refresh_after)
                rewrites.append(
                    {
                        "table": str(table.Name),
                        "partition": str(partition.Name),
                        "previous_expression_length": len(expression),
                        "expression_length": len(new_expression),
                        "preview_after": redact_sensitive_data(new_expression[:200]),
                        "status": "planned" if dry_run else "rewritten",
                    }
                )
        return {
            "parameter_name": parameter_name,
            "parameter_expression": parameter_expression,
            "parameter_action": named_expression_action,
            "file_path": canonical_path,
            "dry_run": dry_run,
            "refresh_requested": refresh_after and not dry_run,
            "rewrites": rewrites,
            "rewrite_count": len(rewrites),
        }

    if dry_run:
        payload = manager.run_read(
            "parameterize_data_source_dry",
            lambda state: _act(state, state.database, state.database.Model),
        )
        return ok(
            f"Dry run: parameter '{parameter_name}' would be {payload['parameter_action']}, "
            f"{payload['rewrite_count']} partition(s) would be rewritten.",
            **payload,
            connection=manager.run_read("snapshot", lambda state: state.snapshot()),
        )
    payload = manager.execute_write("parameterize_data_source", _act)
    return ok(
        f"Parameter '{parameter_name}' {payload['parameter_action']}, "
        f"{payload['rewrite_count']} partition(s) rewritten.",
        parameter_name=payload["parameter_name"],
        parameter_action=payload["parameter_action"],
        file_path=payload["file_path"],
        rewrites=payload["rewrites"],
        rewrite_count=payload["rewrite_count"],
        refresh_requested=payload["refresh_requested"],
        save_result=payload["save_result"],
        persistence=payload.get("persistence"),
        connection=payload["connection"],
    )


def pbi_relocate_data_source_tool(
    manager: Any,
    *,
    old_path: str,
    new_path: str,
    case_sensitive: bool = False,
    dry_run: bool = False,
    refresh_after: bool = False,
) -> dict[str, Any]:
    """Bulk-rewrite a hardcoded file or folder path inside every M partition.

    Use when a workbook moves and queries break with DataSource.NotFound. Looks
    for ``old_path`` as a substring of each M-source partition expression and
    replaces it with ``new_path``. Calculated/query-source partitions are
    skipped. ``dry_run=True`` returns the planned changes without writing.
    """
    if not old_path or not new_path:
        raise PowerBIValidationError("old_path and new_path must both be non-empty.", details={"old_path": old_path, "new_path": new_path})
    if not dry_run:
        # Only validate that the new path resolves when actually rewriting; allows dry-run from
        # any host and avoids resolving paths that may live outside the allowlist when probing.
        resolved_new = resolve_local_path(new_path, must_exist=False)
    else:
        resolved_new = new_path
    needle = old_path if case_sensitive else old_path.casefold()

    def _act(state: Any, database: Any, model: Any) -> dict[str, Any]:
        plan: list[dict[str, Any]] = []
        for table in model.Tables:
            if bool(getattr(table, "IsHidden", False)):
                # still inspect — sometimes hidden internal tables hold the source.
                pass
            for partition in table.Partitions:
                source_type = _source_type_token(partition)
                if source_type != "m":
                    continue
                expression = _partition_expression(partition)
                haystack = expression if case_sensitive else expression.casefold()
                if needle not in haystack:
                    continue
                if case_sensitive:
                    new_expression = expression.replace(old_path, new_path)
                else:
                    # Preserve original casing of non-matching segments via regex w/ case-insensitive flag.
                    import re
                    new_expression = re.sub(re.escape(old_path), lambda _m: new_path, expression, flags=re.IGNORECASE)
                if new_expression == expression:
                    continue
                entry: dict[str, Any] = {
                    "table": str(table.Name),
                    "partition": str(partition.Name),
                    "previous_expression_length": len(expression),
                    "expression_length": len(new_expression),
                    "occurrences": expression.count(old_path) if case_sensitive else len(re.findall(re.escape(old_path), expression, flags=re.IGNORECASE)),
                    "preview_before": redact_sensitive_data(expression[:200]),
                    "preview_after": redact_sensitive_data(new_expression[:200]),
                }
                # Validate the rewritten M up front, even on dry_run, so the
                # caller sees syntax/security errors during preview instead of
                # only when they later commit the change.
                try:
                    _validate_m_expression(new_expression)
                    entry["validation"] = "ok"
                except PowerBIValidationError as exc:
                    entry["validation"] = "invalid"
                    entry["validation_error"] = getattr(exc, "message", str(exc))
                    if not dry_run:
                        # Surface the original exception path on the live write.
                        raise
                if not dry_run:
                    _set_partition_m_expression(manager, database, partition, new_expression)
                    _request_refresh(manager, table, refresh_after)
                    entry["status"] = "rewritten"
                else:
                    entry["status"] = "planned"
                plan.append(entry)
        return {
            "old_path": old_path,
            "new_path": str(resolved_new),
            "case_sensitive": case_sensitive,
            "dry_run": dry_run,
            "rewritten": [item for item in plan if item["status"] == "rewritten"],
            "planned": [item for item in plan if item["status"] == "planned"],
            "match_count": len(plan),
            "refresh_requested": refresh_after and not dry_run,
        }

    if dry_run:
        payload = manager.run_read("relocate_data_source_dry", lambda state: _act(state, state.database, state.database.Model))
        return ok(
            f"Dry run: {payload['match_count']} partitions would be rewritten.",
            **payload,
            connection=manager.run_read("snapshot", lambda state: state.snapshot()),
        )
    payload = manager.execute_write("relocate_data_source", _act)
    return ok(
        f"Relocated data source in {payload['match_count']} partitions.",
        old_path=payload["old_path"],
        new_path=payload["new_path"],
        rewritten=payload["rewritten"],
        match_count=payload["match_count"],
        refresh_requested=payload["refresh_requested"],
        save_result=payload["save_result"],
        persistence=payload.get("persistence"),
        connection=payload["connection"],
    )


def pbi_import_excel_workbook_tool(
    manager: Any,
    *,
    excel_path: str,
    sheet_table_map: dict[str, str] | None = None,
    promote_headers: bool = True,
    refresh_after: bool = True,
) -> dict[str, Any]:
    """Import an Excel workbook into Power BI tables in one call."""
    return pbi_bulk_import_excel_tool(
        manager,
        excel_path=excel_path,
        sheet_table_map=sheet_table_map,
        promote_headers=promote_headers,
        refresh_after=refresh_after,
    )


__all__ = [
    "_build_csv_m",
    "_build_excel_m",
    "_build_folder_m",
    "_validate_m_expression",
    "pbi_bulk_import_excel_tool",
    "pbi_create_csv_import_query_tool",
    "pbi_create_folder_import_query_tool",
    "pbi_create_import_query_tool",
    "pbi_get_power_query_tool",
    "pbi_import_excel_workbook_tool",
    "pbi_list_power_queries_tool",
    "pbi_parameterize_data_source_tool",
    "pbi_relocate_data_source_tool",
    "pbi_set_power_query_tool",
]
