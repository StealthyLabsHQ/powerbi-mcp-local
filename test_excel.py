"""Standalone tests for Excel workbook tools."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from tools.excel import (
    OPENPYXL_AVAILABLE,
    excel_auto_width_tool,
    excel_create_sheet_tool,
    excel_create_workbook_tool,
    excel_delete_sheet_tool,
    excel_format_range_tool,
    excel_read_cell_tool,
    excel_read_sheet_tool,
    excel_search_tool,
    excel_to_pbi_check_tool,
    excel_workbook_info_tool,
    excel_write_cell_tool,
    excel_write_range_tool,
)

if OPENPYXL_AVAILABLE:
    from openpyxl import load_workbook


class _FakeManager:
    def __init__(self) -> None:
        sales = SimpleNamespace(
            Name="Sales",
            Columns=[
                SimpleNamespace(Name="Product", DataType="String"),
                SimpleNamespace(Name="Qty", DataType="Int64"),
                SimpleNamespace(Name="SaleDate", DataType="DateTime"),
            ],
        )
        inventory = SimpleNamespace(
            Name="Inventory",
            Columns=[
                SimpleNamespace(Name="Sku", DataType="String"),
                SimpleNamespace(Name="Stock", DataType="Int64"),
            ],
        )
        self._state = SimpleNamespace(
            database=SimpleNamespace(Model=SimpleNamespace(Tables=[sales, inventory])),
            snapshot=lambda: {"connected": True, "port": 55555, "database": "UnitTest"},
        )

    def run_read(self, _operation: str, reader):
        return reader(self._state)


@unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is required for Excel tool tests")
class ExcelToolTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.workbook_path = self.root / "sample.xlsx"
        created = excel_create_workbook_tool(str(self.workbook_path), sheets=["Sales", "Inventory", "Notes"])
        self.assertTrue(created["ok"], created)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_read_write_and_pipeline(self) -> None:
        wrote = excel_write_range_tool(
            str(self.workbook_path),
            "Sales",
            "A1",
            [["Product", "Qty", "SaleDate"], ["Helmet", 3, "2026-04-16"], ["Gloves", 5, "2026-04-17"]],
        )
        self.assertTrue(wrote["ok"], wrote)

        excel_write_cell_tool(str(self.workbook_path), "Notes", "A1", "Margin =")
        excel_write_cell_tool(str(self.workbook_path), "Notes", "A2", "=SUM(1,2)")

        sheet = excel_read_sheet_tool(str(self.workbook_path), "Sales")
        self.assertTrue(sheet["ok"], sheet)
        self.assertEqual(sheet["headers"], ["Product", "Qty", "SaleDate"])
        self.assertEqual(sheet["rows"][0], ["Helmet", 3, "2026-04-16"])
        self.assertEqual(sheet["returned_rows"], 2)

        cell = excel_read_cell_tool(str(self.workbook_path), "Notes", "A2")
        self.assertTrue(cell["ok"], cell)
        self.assertEqual(cell["formula"], "=SUM(1,2)")

        search = excel_search_tool(str(self.workbook_path), "helmet")
        self.assertTrue(search["ok"], search)
        self.assertEqual(search["results"][0]["sheet"], "Sales")

        info = excel_workbook_info_tool(str(self.workbook_path))
        self.assertTrue(info["ok"], info)
        self.assertEqual(len(info["sheets"]), 3)

        check = excel_to_pbi_check_tool(str(self.workbook_path), _FakeManager())
        self.assertTrue(check["ok"], check)
        self.assertEqual(check["matches"][0]["table"], "Sales")
        self.assertIn("Product", check["matches"][0]["matched_columns"])

    def test_sheet_management_and_formatting(self) -> None:
        created = excel_create_sheet_tool(str(self.workbook_path), "Scratch", position=1)
        self.assertTrue(created["ok"], created)

        excel_write_range_tool(
            str(self.workbook_path),
            "Scratch",
            "A1",
            [["Metric", "Value"], ["Revenue", 1200], ["Cost", 800]],
        )
        formatted = excel_format_range_tool(
            str(self.workbook_path),
            "Scratch",
            "A1:B1",
            {"bold": True, "fill_color": "D9EAF7", "font_color": "1F4E78", "alignment": "center"},
        )
        self.assertTrue(formatted["ok"], formatted)

        auto_width = excel_auto_width_tool(str(self.workbook_path), "Scratch")
        self.assertTrue(auto_width["ok"], auto_width)

        workbook = load_workbook(self.workbook_path)
        worksheet = workbook["Scratch"]
        self.assertTrue(worksheet["A1"].font.bold)
        self.assertEqual(worksheet["A1"].fill.fgColor.rgb, "FFD9EAF7")
        self.assertGreater(worksheet.column_dimensions["A"].width, 8)
        workbook.close()

        deleted = excel_delete_sheet_tool(str(self.workbook_path), "Scratch")
        self.assertTrue(deleted["ok"], deleted)

    def test_error_handling(self) -> None:
        missing = excel_read_sheet_tool(str(self.root / "missing.xlsx"), "Sales")
        self.assertFalse(missing["ok"])
        self.assertEqual(missing["error"]["code"], "excel_file_not_found")

        with patch("openpyxl.workbook.workbook.Workbook.save", side_effect=PermissionError("locked")):
            locked = excel_write_cell_tool(str(self.workbook_path), "Sales", "A1", "Blocked")
        self.assertFalse(locked["ok"])
        self.assertEqual(locked["error"]["code"], "excel_file_locked")


if __name__ == "__main__":
    unittest.main(verbosity=2)
