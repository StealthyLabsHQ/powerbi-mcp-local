"""Excel workbook tools for the Power BI MCP server."""

from __future__ import annotations

import difflib
import os
from contextlib import contextmanager
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Border, PatternFill, Side
    from openpyxl.utils import get_column_letter, range_boundaries

    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - dependency is optional until installed
    OPENPYXL_AVAILABLE = False
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]
    Border = PatternFill = Side = None  # type: ignore[assignment]
    get_column_letter = range_boundaries = None  # type: ignore[assignment]

from pbi_connection import PowerBIError, error_payload, normalize_token, ok, serialize_value

LARGE_FILE_BYTES = 10 * 1024 * 1024
DEFAULT_READ_LIMIT = 500
MAX_SEARCH_RESULTS = 500

class ExcelToolError(PowerBIError): code = "excel_error"
class ExcelDependencyError(ExcelToolError): code = "excel_dependency_error"
class ExcelFileNotFoundError(ExcelToolError): code = "excel_file_not_found"
class ExcelSheetNotFoundError(ExcelToolError): code = "excel_sheet_not_found"
class ExcelFileLockedError(ExcelToolError): code = "excel_file_locked"
class ExcelValidationError(ExcelToolError): code = "excel_validation_error"
def _run(callback: Any, *args: Any, **kwargs: Any) -> dict[str, Any]:
    try:
        return callback(*args, **kwargs)
    except Exception as exc:
        return error_payload(exc)

def _ensure_openpyxl() -> None:
    if not OPENPYXL_AVAILABLE:
        raise ExcelDependencyError("openpyxl is not installed. Install it with 'pip install openpyxl'.")

def _resolve_path(file_path: str) -> Path:
    text = str(file_path).strip()
    if not text:
        raise ExcelValidationError("file_path cannot be empty.")
    normalized = os.path.normpath(text.replace("\\", os.sep))
    path = Path(normalized).expanduser()
    if not path.is_absolute():
        path = Path.cwd() / path
    return path.resolve(strict=False)

def _require_file(path: Path) -> None:
    if not path.exists(): raise ExcelFileNotFoundError(f"Workbook not found: {path}", details={"path": str(path)})

def _streaming(path: Path) -> bool: return path.exists() and path.stat().st_size > LARGE_FILE_BYTES

@contextmanager
def _open_workbook(
    file_path: str,
    *,
    read_only: bool | None = None,
    data_only: bool = False,
    for_write: bool = False,
) -> Iterator[tuple[Any, Path, bool]]:
    _ensure_openpyxl()
    path = _resolve_path(file_path)
    if not for_write:
        _require_file(path)
    effective_read_only = False if for_write else _streaming(path) if read_only is None else read_only
    try:
        workbook = load_workbook(path, read_only=effective_read_only, data_only=data_only)
    except FileNotFoundError as exc:
        raise ExcelFileNotFoundError(f"Workbook not found: {path}", details={"path": str(path)}) from exc
    except PermissionError as exc:
        raise ExcelFileLockedError("File locked by Excel, close it first.", details={"path": str(path)}) from exc
    try:
        yield workbook, path, effective_read_only
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()

def _save_workbook(workbook: Any, path: Path) -> None:
    try:
        workbook.save(path)
    except PermissionError as exc:
        raise ExcelFileLockedError("File locked by Excel, close it first.", details={"path": str(path)}) from exc

def _sheet(workbook: Any, name: str) -> Any:
    if name not in workbook.sheetnames:
        raise ExcelSheetNotFoundError(
            f"Sheet '{name}' not found.",
            details={"sheet": name, "available_sheets": list(workbook.sheetnames)},
        )
    return workbook[name]

def _value_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, (date, datetime)):
        return "datetime"
    return "string"

def _header_name(value: Any, index: int) -> str:
    text = "" if value is None else str(serialize_value(value)).strip()
    return text or f"Column{index}"

def _normalize_color(value: str) -> str:
    token = value.strip().lstrip("#")
    if len(token) == 6:
        return "FF" + token.upper()
    if len(token) == 8:
        return token.upper()
    raise ExcelValidationError("Color values must be 6 or 8 hex characters.", details={"value": value})

def _sheet_summary(sheet: Any) -> dict[str, Any]:
    row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    has_data = bool(row and any(cell is not None for cell in row)) or sheet.max_row > 1
    try:
        dimensions = str(sheet.calculate_dimension())
    except Exception:
        dimensions = None
    return {
        "name": sheet.title,
        "rows": sheet.max_row,
        "columns": sheet.max_column,
        "has_data": has_data,
        "dimensions": dimensions,
    }

def _infer_excel_type(rows: list[list[Any]], index: int) -> str:
    for row in rows:
        if index < len(row) and row[index] is not None:
            return _value_type(row[index])
    return "unknown"

def _type_issue(excel_type: str, pbi_type: str) -> bool:
    token = pbi_type.casefold()
    if excel_type == "unknown":
        return False
    if excel_type == "number":
        return any(part in token for part in ("string", "text"))
    if excel_type == "string":
        return any(part in token for part in ("int", "decimal", "double", "number", "whole"))
    if excel_type == "datetime":
        return not any(part in token for part in ("date", "time"))
    if excel_type == "boolean":
        return "bool" not in token
    return False

def excel_list_sheets_tool(file_path: str) -> dict[str, Any]:
    def _impl() -> dict[str, Any]:
        with _open_workbook(file_path, data_only=True) as (workbook, path, read_only):
            return ok(
                "Workbook sheets listed successfully.",
                file_path=str(path),
                read_only=read_only,
                sheets=[_sheet_summary(sheet) for sheet in workbook.worksheets],
            )
    return _run(_impl)

def excel_read_sheet_tool(
    file_path: str,
    sheet: str,
    range: str | None = None,
    limit: int = DEFAULT_READ_LIMIT,
) -> dict[str, Any]:
    def _impl() -> dict[str, Any]:
        if limit < 1:
            raise ExcelValidationError("limit must be >= 1.", details={"limit": limit})
        with _open_workbook(file_path, data_only=True) as (workbook, path, read_only):
            worksheet = _sheet(workbook, sheet)
            bounds = range_boundaries(range) if range else (1, 1, worksheet.max_column, worksheet.max_row)
            rows_iter = worksheet.iter_rows(
                min_col=bounds[0],
                min_row=bounds[1],
                max_col=bounds[2],
                max_row=bounds[3],
                values_only=True,
            )
            first_row = next(rows_iter, None)
            if first_row is None:
                return ok(
                    "Sheet range is empty.",
                    file_path=str(path),
                    sheet=sheet,
                    range=range,
                    read_only=read_only,
                    headers=[],
                    rows=[],
                    total_rows=0,
                    returned_rows=0,
                    truncated=False,
                )
            headers = [_header_name(value, index) for index, value in enumerate(first_row, start=1)]
            rows, total_rows = [], 0
            for row in rows_iter:
                total_rows += 1
                if len(rows) < limit:
                    rows.append([serialize_value(value) for value in row])
            return ok(
                "Sheet read successfully.",
                file_path=str(path),
                sheet=sheet,
                range=range,
                read_only=read_only,
                headers=headers,
                rows=rows,
                total_rows=total_rows,
                returned_rows=len(rows),
                truncated=total_rows > len(rows),
            )
    return _run(_impl)

def excel_read_cell_tool(file_path: str, sheet: str, cell: str) -> dict[str, Any]:
    def _impl() -> dict[str, Any]:
        with _open_workbook(file_path, data_only=False) as (formula_wb, path, read_only):
            formula_cell = _sheet(formula_wb, sheet)[cell]
            formula = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None
            number_format = formula_cell.number_format
        with _open_workbook(str(path), read_only=read_only, data_only=True) as (value_wb, _, _):
            value = _sheet(value_wb, sheet)[cell].value
        return ok(
            "Cell read successfully.",
            file_path=str(path),
            sheet=sheet,
            cell=cell,
            value=serialize_value(value),
            type=_value_type(value),
            format=number_format,
            formula=formula,
        )
    return _run(_impl)

def excel_search_tool(file_path: str, query: str, sheet: str | None = None) -> dict[str, Any]:
    def _impl() -> dict[str, Any]:
        needle = query.casefold().strip()
        if not needle:
            raise ExcelValidationError("query cannot be empty.")
        with _open_workbook(file_path, data_only=True) as (workbook, path, read_only):
            results, truncated = [], False
            for sheet_name in [sheet] if sheet else workbook.sheetnames:
                worksheet = _sheet(workbook, sheet_name)
                for row in worksheet.iter_rows():
                    for cell in row:
                        value = cell.value
                        if value is not None and needle in str(serialize_value(value)).casefold():
                            if len(results) >= MAX_SEARCH_RESULTS:
                                truncated = True
                                break
                            results.append({"sheet": sheet_name, "cell": cell.coordinate, "value": serialize_value(value)})
                    if truncated:
                        break
                if truncated:
                    break
            return ok(
                "Search completed successfully.",
                file_path=str(path),
                query=query,
                read_only=read_only,
                results=results,
                total_matches=len(results),
                truncated=truncated,
            )
    return _run(_impl)

def excel_write_cell_tool(file_path: str, sheet: str, cell: str, value: Any, format: str = "") -> dict[str, Any]:
    def _impl() -> dict[str, Any]:
        with _open_workbook(file_path, for_write=True) as (workbook, path, _):
            target = _sheet(workbook, sheet)[cell]
            target.value = value
            if format:
                target.number_format = format
            _save_workbook(workbook, path)
            return ok("Cell written successfully.", file_path=str(path), sheet=sheet, cell=cell, value=serialize_value(value), format=format or None)
    return _run(_impl)

def excel_write_range_tool(file_path: str, sheet: str, start_cell: str, data: list[list[Any]]) -> dict[str, Any]:
    def _impl() -> dict[str, Any]:
        if not data or any(not isinstance(row, (list, tuple)) for row in data):
            raise ExcelValidationError("data must be a non-empty 2D array.")
        with _open_workbook(file_path, for_write=True) as (workbook, path, _):
            worksheet = _sheet(workbook, sheet)
            anchor = worksheet[start_cell]
            width = 0
            for row_offset, row in enumerate(data):
                width = max(width, len(row))
                for col_offset, value in enumerate(row):
                    worksheet.cell(row=anchor.row + row_offset, column=anchor.column + col_offset, value=value)
            _save_workbook(workbook, path)
            return ok(
                "Range written successfully.",
                file_path=str(path),
                sheet=sheet,
                start_cell=start_cell,
                rows_written=len(data),
                columns_written=width,
            )
    return _run(_impl)

def excel_create_sheet_tool(file_path: str, name: str, position: int | None = None) -> dict[str, Any]:
    def _impl() -> dict[str, Any]:
        with _open_workbook(file_path, for_write=True) as (workbook, path, _):
            if name in workbook.sheetnames:
                raise ExcelValidationError("Sheet already exists.", details={"sheet": name})
            workbook.create_sheet(title=name, index=position if position is not None else len(workbook.sheetnames))
            _save_workbook(workbook, path)
            return ok("Sheet created successfully.", file_path=str(path), sheet=name, position=position)
    return _run(_impl)

def excel_delete_sheet_tool(file_path: str, name: str) -> dict[str, Any]:
    def _impl() -> dict[str, Any]:
        with _open_workbook(file_path, for_write=True) as (workbook, path, _):
            worksheet = _sheet(workbook, name)
            if len(workbook.sheetnames) == 1:
                raise ExcelValidationError("Cannot delete the last remaining sheet.", details={"sheet": name})
            workbook.remove(worksheet)
            _save_workbook(workbook, path)
            return ok("Sheet deleted successfully.", file_path=str(path), sheet=name)
    return _run(_impl)

def excel_format_range_tool(file_path: str, sheet: str, range: str, format: dict[str, Any]) -> dict[str, Any]:
    def _impl() -> dict[str, Any]:
        if not isinstance(format, dict) or not format:
            raise ExcelValidationError("format must be a non-empty object.")
        with _open_workbook(file_path, for_write=True) as (workbook, path, _):
            worksheet = _sheet(workbook, sheet)
            bounds = range_boundaries(range)
            for row in worksheet.iter_rows(min_col=bounds[0], min_row=bounds[1], max_col=bounds[2], max_row=bounds[3]):
                for cell in row:
                    font = copy(cell.font)
                    alignment = copy(cell.alignment)
                    if "bold" in format:
                        font.bold = bool(format["bold"])
                    if "italic" in format:
                        font.italic = bool(format["italic"])
                    if "font_size" in format:
                        font.size = format["font_size"]
                    if "font_color" in format:
                        font.color = _normalize_color(str(format["font_color"]))
                    cell.font = font
                    if "fill_color" in format:
                        cell.fill = PatternFill(fill_type="solid", fgColor=_normalize_color(str(format["fill_color"])))
                    if "number_format" in format:
                        cell.number_format = str(format["number_format"])
                    if "alignment" in format:
                        alignment.horizontal = str(format["alignment"])
                        cell.alignment = alignment
                    if "border" in format:
                        side = Side(style=str(format["border"]))
                        cell.border = Border(left=side, right=side, top=side, bottom=side)
            _save_workbook(workbook, path)
            return ok("Range formatted successfully.", file_path=str(path), sheet=sheet, range=range, format=format)
    return _run(_impl)

def excel_auto_width_tool(file_path: str, sheet: str) -> dict[str, Any]:
    def _impl() -> dict[str, Any]:
        with _open_workbook(file_path, for_write=True) as (workbook, path, _):
            worksheet = _sheet(workbook, sheet)
            updated = []
            for index in range(1, worksheet.max_column + 1):
                letter = get_column_letter(index)
                values = (len(str(cell.value)) if cell.value is not None else 0 for cell in worksheet[letter])
                worksheet.column_dimensions[letter].width = min(max(max(values, default=0) + 2, 8), 60)
                updated.append(letter)
            _save_workbook(workbook, path)
            return ok("Column widths updated successfully.", file_path=str(path), sheet=sheet, columns=updated)
    return _run(_impl)

def excel_create_workbook_tool(file_path: str, sheets: list[str] | None = None) -> dict[str, Any]:
    def _impl() -> dict[str, Any]:
        _ensure_openpyxl()
        path = _resolve_path(file_path)
        if path.exists():
            raise ExcelValidationError("Workbook already exists.", details={"path": str(path)})
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        names = [str(name).strip() for name in (sheets or []) if str(name).strip()]
        if names:
            if len(names) != len(set(names)):
                raise ExcelValidationError("Sheet names must be unique.", details={"sheets": names})
            workbook.active.title = names[0]
            for sheet_name in names[1:]:
                workbook.create_sheet(title=sheet_name)
        created_sheets = names or [workbook.active.title]
        _save_workbook(workbook, path)
        workbook.close()
        return ok("Workbook created successfully.", file_path=str(path), sheets=created_sheets)
    return _run(_impl)

def excel_workbook_info_tool(file_path: str) -> dict[str, Any]:
    def _impl() -> dict[str, Any]:
        with _open_workbook(file_path, data_only=True) as (workbook, path, read_only):
            properties = workbook.properties
            names = []
            for name, value in getattr(workbook, "defined_names", {}).items():
                names.append({"name": str(name), "value": str(value)})
            return ok(
                "Workbook info collected successfully.",
                file_path=str(path),
                read_only=read_only,
                sheets=[_sheet_summary(sheet) for sheet in workbook.worksheets],
                named_ranges=names,
                properties={
                    "title": serialize_value(properties.title),
                    "creator": serialize_value(properties.creator),
                    "created": serialize_value(properties.created),
                    "modified": serialize_value(properties.modified),
                    "last_modified_by": serialize_value(properties.lastModifiedBy),
                },
            )
    return _run(_impl)

def excel_to_pbi_check_tool(file_path: str, manager: Any) -> dict[str, Any]:
    def _impl() -> dict[str, Any]:
        with _open_workbook(file_path, data_only=True) as (workbook, path, read_only):
            pbi_payload = manager.run_read(
                "excel_to_pbi_check",
                lambda state: {
                    "connection": state.snapshot(),
                    "tables": [
                        {
                            "name": str(table.Name),
                            "columns": [
                                {"name": str(column.Name), "data_type": str(getattr(column, "DataType", ""))}
                                for column in table.Columns
                            ],
                        }
                        for table in state.database.Model.Tables
                    ],
                },
            )
            tables = pbi_payload["tables"]
            lookup = {normalize_token(item["name"]): item for item in tables}
            matches, mismatches, suggestions = [], [], []
            for worksheet in workbook.worksheets:
                if not _sheet_summary(worksheet)["has_data"]:
                    continue
                table = lookup.get(normalize_token(worksheet.title))
                if table is None:
                    close = difflib.get_close_matches(worksheet.title, [item["name"] for item in tables], n=3)
                    mismatches.append({"sheet": worksheet.title, "issue": "no_matching_table"})
                    if close:
                        suggestions.append({"sheet": worksheet.title, "suggested_tables": close})
                    continue
                rows = list(worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 21), values_only=True))
                headers = [_header_name(value, index) for index, value in enumerate(rows[0], start=1)] if rows else []
                excel_lookup = {normalize_token(name): name for name in headers}
                model_lookup = {normalize_token(item["name"]): item for item in table["columns"]}
                type_issues = []
                for index, header in enumerate(headers):
                    model_column = model_lookup.get(normalize_token(header))
                    if model_column is None:
                        continue
                    excel_type = _infer_excel_type([list(row) for row in rows[1:]], index)
                    if _type_issue(excel_type, str(model_column["data_type"])):
                        type_issues.append({"column": header, "excel_type": excel_type, "pbi_type": model_column["data_type"]})
                matches.append(
                    {
                        "sheet": worksheet.title,
                        "table": table["name"],
                        "matched_columns": [model_lookup[key]["name"] for key in model_lookup if key in excel_lookup],
                        "missing_in_excel": [item["name"] for key, item in model_lookup.items() if key not in excel_lookup],
                        "extra_in_excel": [name for key, name in excel_lookup.items() if key not in model_lookup],
                        "type_issues": type_issues,
                    }
                )
            return ok(
                "Excel workbook checked against the Power BI model successfully.",
                file_path=str(path),
                read_only=read_only,
                connection=pbi_payload["connection"],
                matches=matches,
                mismatches=mismatches,
                suggestions=suggestions,
            )
    return _run(_impl)

__all__ = [
    "excel_auto_width_tool", "excel_create_sheet_tool", "excel_create_workbook_tool", "excel_delete_sheet_tool",
    "excel_format_range_tool", "excel_list_sheets_tool", "excel_read_cell_tool", "excel_read_sheet_tool",
    "excel_search_tool", "excel_to_pbi_check_tool", "excel_workbook_info_tool", "excel_write_cell_tool",
    "excel_write_range_tool",
]
